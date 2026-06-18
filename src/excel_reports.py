from openpyxl import Workbook

from openpyxl import Workbook

def generate_sales_report(df):

    wb = Workbook()

    ws = wb.active

    ws.title = "Sales Report"

    ws.append([
        "Date",
        "Product",
        "Units Sold"
    ])

    for row in df.itertuples(index=False):

        ws.append(list(row))

    wb.save(
        "reports/sales_report.xlsx"
    )

    print(
        "Sales Report Generated!"
    )
def generate_forecast_report(
    forecast_results
):

    wb = Workbook()

    ws = wb.active

    ws.title = "Forecast Report"

    ws.append([
        "Product",
        "Predicted Demand"
    ])

    for product, demand in forecast_results.items():

        ws.append([
            product,
            demand
        ])

    wb.save(
        "reports/forecast_report.xlsx"
    )

    print(
        "Forecast Report Generated!"
    )
def generate_inventory_report(
    inventory_report
):

    wb = Workbook()

    ws = wb.active

    ws.title = "Inventory Report"

    ws.append([
        "Product",
        "Current Stock",
        "Minimum Stock",
        "Status"
    ])

    for product, data in inventory_report.items():

        ws.append([
            product,
            data["Current Stock"],
            data["Minimum Stock"],
            data["Status"]
        ])

    wb.save(
        "reports/inventory_report.xlsx"
    )

    print(
        "Inventory Report Generated!"
    )
def generate_purchase_order_excel(
    risk_report
):

    wb = Workbook()

    ws = wb.active

    ws.title = "Purchase Orders"

    ws.append([
        "Product",
        "Current Stock",
        "Predicted Demand",
        "Required Purchase"
    ])

    for product, data in risk_report.items():

        if data["Shortage"] > 0:

            ws.append([
                product,
                data["Current Stock"],
                data["Predicted Demand"],
                data["Shortage"]
            ])

    wb.save(
        "reports/purchase_order.xlsx"
    )

    print(
        "Purchase Order Excel Generated!"
    )